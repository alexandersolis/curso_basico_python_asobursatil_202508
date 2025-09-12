#Recuerda instalar openpyxl: pip install openpyxl
import openpyxl

# Dcumentacion oficial openpyxl: Estilos
# https://openpyxl.readthedocs.io/en/stable/styles.html
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# Crear un nuevo libro de Excel
libro = openpyxl.Workbook()

# Seleccionar una hoja
hoja = libro.active

# Escribir datos en celdas
hoja['A1'] = 'Nombre'
hoja['B1'] = 'Edad'
hoja['A2'] = 'Juan'
hoja['B2'] = 25
hoja['A3'] = 'Pedro'
hoja['B3'] = 30
hoja['A4'] = 'Carlos'
hoja['B4'] = 20

# Asi puedo definir una formula en una celda
celda = hoja.cell(row=5, column=2)  
celda.value = "=SUM(B1:B4)"  
#Poner en negrita la celda
celda.font = Font(bold=True)  

# Combinar celdas
hoja.merge_cells('A7:C7')  
celda = hoja.cell(row=7, column=1)  
celda.value = 'Acciones y Valores'  

# Alinear celdas
celda.alignment = Alignment(horizontal='center', vertical='center')  

# Dar estilo a celdas
fuente = Font(name='Arial', size=14, bold=True, color="FF0000")
celda = hoja['A1']
celda.font = fuente

# Guardar el libro como un nuevo archivo
libro.save('tema_7_archivos/09_crear_archivo_excel_con_estilos.xlsx')
